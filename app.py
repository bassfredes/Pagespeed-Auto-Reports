import time
import datetime
import urllib.parse
import openpyxl
from openpyxl import Workbook
import logging
import logging.handlers
from concurrent.futures import ProcessPoolExecutor, as_completed
import random
import multiprocessing
from collections import Counter
from seleniumbase import BaseCase
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException

# Cola global de logs para procesos
LOG_QUEUE = multiprocessing.Queue(-1)

def init_child_logger(queue):
    """
    Función inicializadora que se ejecuta en cada proceso hijo.
    Configura el logger para enviar los mensajes a la cola.
    """
    root_logger = logging.getLogger()
    for h in root_logger.handlers[:]:
        root_logger.removeHandler(h)
    queue_handler = logging.handlers.QueueHandler(queue)
    root_logger.addHandler(queue_handler)
    root_logger.setLevel(logging.INFO)

# Configuración del logger en el proceso principal
logger = logging.getLogger(__name__)
logger.setLevel(logging.INFO)
file_handler = logging.FileHandler("app_log.log", mode="w")
formatter = logging.Formatter("%(asctime)s [%(levelname)s] %(processName)s - %(message)s", "%Y-%m-%d %H:%M:%S")
file_handler.setFormatter(formatter)
console_handler = logging.StreamHandler()
console_handler.setFormatter(formatter)
listener = logging.handlers.QueueListener(LOG_QUEUE, file_handler, console_handler)
listener.start()

class PagespeedAnalyzer(BaseCase):
    INPUT_EXCEL = 'sites.xlsx'   # Archivo de entrada con la columna "url"
    OUTPUT_EXCEL = 'resultados.xlsx'
    REPORT_COUNT = 5             # Número de análisis por URL

    def read_urls(self):
        """Lee las URLs desde un archivo Excel."""
        wb = openpyxl.load_workbook(self.INPUT_EXCEL)
        sheet = wb.active
        urls = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0]:
                urls.append(row[0])
        return urls
      
    @staticmethod
    def add_v_parameter(url, timestamp):
        """Agrega el parámetro 'v' respetando la query y el fragmento de la URL."""
        parsed = urllib.parse.urlsplit(url)
        if parsed.query:
            new_query = f"{parsed.query}&v={timestamp}"
        else:
            new_query = f"v={timestamp}"
        new_url = urllib.parse.urlunsplit((
            parsed.scheme,
            parsed.netloc,
            parsed.path,
            new_query,
            parsed.fragment
        ))
        return new_url

    def _run_analysis_internal(self, test_url):
        logger.info(f"Iniciando análisis de URL: {test_url}")
        self.open("https://pagespeed.web.dev/")
        try:
            input_element = WebDriverWait(self.driver, 10).until(
                EC.presence_of_element_located((By.ID, "i4"))
            )
            input_element.clear()
            input_element.send_keys(test_url)
            
            analyze_button = WebDriverWait(self.driver, 10).until(
                EC.element_to_be_clickable((By.XPATH, "//button[.//span[contains(text(),'Analyze')]]"))
            )
            analyze_button.click()
            
            try:
                error_element = WebDriverWait(self.driver, 30).until(
                    EC.presence_of_element_located((By.XPATH, "//div[@class='RDG5zd']"))
                )
                if error_element:
                    raise Exception("RPC::DEADLINE_EXCEEDED: context deadline exceeded")
            except TimeoutException:
                pass

            performance_element = WebDriverWait(self.driver, 120).until(
                EC.presence_of_element_located((
                    By.XPATH,
                    "//div[@id='performance']//*[local-name()='text' and contains(@class, 'lh-exp-gauge__percentage')]"
                ))
            )
            time.sleep(2)
            performance_html = self.driver.execute_script("return arguments[0].textContent;", performance_element)
            
            performance = int(performance_html.strip())
            report_url = self.get_current_url()
            logger.info(f"Análisis exitoso: {performance}%")
        except TimeoutException:
            raise Exception(f"Timeout al analizar {test_url}")
        return performance, report_url

def run_analysis_in_process(base_url):
    """
    Función que se ejecuta en cada proceso hijo.
    Genera, en ese contexto, un valor aleatorio para el parámetro 'v',
    construye la URL de análisis y ejecuta el análisis.
    """
    # Espera aleatoria breve para diferenciar instancias
    time.sleep(random.uniform(0.1, 0.5))
    ts = int(time.time() * 1000) + random.randint(0, 1000000)
    variation = PagespeedAnalyzer.add_v_parameter(base_url, ts)
    logging.info(f"URL generada en proceso: {variation}")
    analyzer = PagespeedAnalyzer()
    analyzer.setUp()  # Inicializa driver y entorno
    try:
        result = analyzer._run_analysis_internal(variation)
    finally:
        analyzer.tearDown()
    return result

def analyze_all_urls_main():
    """
    Función principal que lee las URLs y, para cada URL,
    ejecuta REPORT_COUNT análisis en paralelo (cada uno generando su propia variación).
    Registra los resultados en un archivo Excel y escribe logs.
    """
    analyzer = PagespeedAnalyzer()
    urls = analyzer.read_urls()
    wb = Workbook()
    ws = wb.active
    ws.append(["URL Base", "Reporte", "Performance", "Reporte URL", "Timestamp", "Observaciones"])

    total = len(urls)
    processed = 0
    max_workers = 5

    with ProcessPoolExecutor(max_workers=max_workers, initializer=init_child_logger, initargs=(LOG_QUEUE,)) as executor:
        for url in urls:
            logging.info(f"Procesando URL base: {url}")
            future_map = {}
            results = []  # Guardará (reporte_id, performance, rep_url)

            for rpt in range(PagespeedAnalyzer.REPORT_COUNT):
                logging.info(f"  Creando tarea para Reporte {rpt+1} de {url}")
                future = executor.submit(run_analysis_in_process, url)
                future_map[rpt+1] = future

            for future in as_completed(future_map.values()):
                rpt_id = None
                for key, fut in future_map.items():
                    if fut == future:
                        rpt_id = key
                        break

                observacion = ""
                try:
                    perf, rep_url = future.result()
                    logging.info(f"  Finalizado reporte {rpt_id} para {url}: {perf}%")
                except Exception as e:
                    perf, rep_url = "Error", ""
                    observacion = str(e)
                    logging.error(f"  Error en reporte {rpt_id} para {url}: {observacion}")

                ws.append([url, rpt_id, perf, rep_url, datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"), observacion])
                
                if isinstance(perf, int):
                    results.append((rpt_id, perf, rep_url))  # Guardamos el reporte, score y URL

            # Procesar resultados para determinar el valor representativo
            if results:
                scores = [r[1] for r in results]  # Lista solo de scores
                score_counts = Counter(scores)  # Cuenta frecuencias
                most_common = score_counts.most_common(1)[0]  # Encuentra la moda

                if most_common[1] > 1:
                    # Si hay moda (repetición), tomamos el más frecuente
                    representative_score = most_common[0]
                else:
                    # Si no hay moda, tomamos la mediana
                    scores.sort()
                    mid = len(scores) // 2
                    representative_score = scores[mid] if len(scores) % 2 != 0 else (scores[mid - 1] + scores[mid]) / 2
                
                # Encontrar el reporte correspondiente al valor representativo
                representative_report = next(r for r in results if r[1] == representative_score)

                ws.append([
                    url, "Resumen", representative_score, representative_report[2], 
                    datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"), 
                    f"Reporte más representativo: {representative_report[0]}"
                ])
            else:
                ws.append([
                    url, "Resumen", "Sin datos", "", 
                    datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"), 
                    "No se pudo obtener performance"
                ])

            processed += 1
            wb.save(PagespeedAnalyzer.OUTPUT_EXCEL)
            logging.info(f"Procesadas {processed} de {total} URL(s). Excel actualizado.")
    logging.info("Proceso completado.")

# Para que pytest detecte la prueba (aunque aquí se ejecuta la función principal)
class TestPagespeed:
    def test_analyze_all_urls(self):
        analyze_all_urls_main()

if __name__ == "__main__":
    import pytest
    print(f"Running {__file__}")
    pytest.main([__file__])
    listener.stop()
